import pandas as pd
from database import get_engine
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image


engine = get_engine()


queries = {

    "Executive Summary": """
        SELECT *
        FROM vw_executive_dashboard
    """,

    "Market Analysis": """
        SELECT *
        FROM vw_market_analysis
    """,

    "Trader Performance": """
        SELECT *
        FROM vw_trader_performance
    """,

    "Renewable Analysis": """
        SELECT *
        FROM vw_renewable_analysis
    """
}


output_file = "reports/Energy_Market_Report_Final.xlsx"


with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

    for sheet_name, query in queries.items():

        df = pd.read_sql(query, engine)

        df.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False
        )


# Formatting
wb = load_workbook(output_file)


for ws in wb:

    ws.freeze_panes = "A2"

    for cell in ws[1]:

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center"
        )


    for column in ws.columns:

        max_length = 0

        column_letter = column[0].column_letter

        for cell in column:

            if cell.value:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        ws.column_dimensions[column_letter].width = max_length + 3


# Add charts

ws = wb["Market Analysis"]

img = Image(
    "charts/market_volume.png"
)

ws.add_image(
    img,
    "G2"
)


ws = wb["Trader Performance"]

img = Image(
    "charts/trader_pnl.png"
)

ws.add_image(
    img,
    "F2"
)


ws = wb["Renewable Analysis"]

img = Image(
    "charts/renewable_mix.png"
)

ws.add_image(
    img,
    "E2"
)


wb.save(output_file)


print("Final Excel report created successfully")